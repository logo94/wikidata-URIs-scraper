from operator import contains
from selenium.webdriver import Firefox
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl
import tkinter as tk
from tkinter import filedialog


url = "https://www.wikidata.org/wiki/Wikidata:Main_Page" # URL della pagina principale di Wikidata

root = tk.Tk()
root.withdraw()
file = filedialog.askopenfilename() # Selezione del foglio di calcolo

fire_driver = GeckoDriverManager().install() # Installazione del driver
driver = Firefox(service=FirefoxService(fire_driver))
driver.maximize_window()
driver.get(url) # Connessione del driver alla pagina principale di Wikidata

wb = openpyxl.load_workbook(file) # Apertura del foglio di calcolo per la lettura e la scrittura
sh = wb.active
m_row = sh.max_row

def title(cell):
    string = sh.cell(row = i, column=2).value # Impostazione del termine presente nel foglio di calcolo come variabile
    return string

def get_uri(i): # Lo scraper carica la pagina principale di Wikidata e segue gli elementi HTML in modo da ricercare il termine desiderato
    driver.get(url)
    txt = title(i)
    driver.find_element(By.NAME, 'search').send_keys(
        txt)
    sleep(1)
    driver.find_element(By.ID, 'searchButton').click()
    sleep(1)

    if driver.find_element(By.XPATH, '/html/body/div[3]/div[3]/div[4]/div[4]/div[2]/ul/li[1]/table/tbody/tr/td[2]/div[1]/a/span/span[1]'):
        driver.find_element(By.XPATH, '/html/body/div[3]/div[3]/div[4]/div[4]/div[2]/ul/li[1]/table/tbody/tr/td[2]/div[1]/a/span/span[1]').click()
        sleep(1)
        abText = driver.current_url # Se l'elemento HTML esiste, l'URL della pagina viene salvato come variabile
        abCell = sh.cell(row = i, column=1)
        abCell.value = "<" + abText + ">" # l'URL viene trascritto all'interno del foglio di calcolo
        wb.save(file) # alvataggio del file
    else:
        pass # In caso di errori il termine viene saltato e si procede con la riga successiva

for i in range(1, m_row + 1): # Per ogni riga del foglio di calcolo viene eseguita l'intera procedura
    try:
        get_uri(i)
    except:
        pass
