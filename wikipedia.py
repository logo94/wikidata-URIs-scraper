from selenium.webdriver import Firefox
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from time import sleep
import openpyxl
import tkinter as tk
from tkinter import filedialog


url = "https://it.wikipedia.org/wiki/Pagina_principale" # URL della pagina principale di Wikipedia Italia

root = tk.Tk()
root.withdraw()
file = filedialog.askopenfilename() # Selezione del foglio di calcolo

fire_driver = GeckoDriverManager().install() # Installazione driver
driver = Firefox(service=FirefoxService(fire_driver))
driver.maximize_window()
driver.get(url) # Caricamento della pagina principale di Wikipedia Italia

wb = openpyxl.load_workbook(file) # Apertura foglio di calcolo per la lettura e la scrittura
sh = wb.active
m_row = sh.max_row

def title(cell):
    string = sh.cell(row = i, column=2).value # Impostazione del termine presente nel foglio di calcolo come variabile
    return string

def get_uri(i): # Lo scraper carica la pagina principale di Wikipedia Italia e segue gli elementi HTML in modo da ricercare il termine desiderato
    driver.get(url)
    txt = title(i)
    driver.find_element(By.NAME, 'search').send_keys(
        txt)
    sleep(1)
    driver.find_element(By.ID, 'searchButton').click()
    sleep(3)
    
    driver.find_element(By.XPATH,
    '/html/body/div[4]/div[2]/nav[3]/div/ul/li[8]/a/span').click()
        
    abText = driver.current_url # Se il termine combacia con la voce Wikipedia, viene seguito l'elemento Wikidata e copiato il relativo URL
    abCell = sh.cell(row = i, column=1)
    abCell.value = "<" + abText + ">" # L'URL viene trascritto all'interno del foglio di calcolo
    wb.save(file) # Salvataggio del file

for i in range(1, m_row + 1): # Per ogni riga del foglio di calcolo viene eseguita l'intera procedura
    try:
        get_uri(i)
    except:
        pass
